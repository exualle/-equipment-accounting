from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.db.utils import IntegrityError
from django.core.exceptions import ValidationError
from django.core import serializers
from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view
from openpyxl import Workbook
import json, datetime
from .models import ItemModel, ItemRelationModel


def index(request):
    return render(request, 'index/index.html')


@csrf_exempt
@api_view(['GET', 'POST'])
def register(request):
    body = json.loads(request.body)
    if body['password'] == body['confirm']:
        try:
            user = User.objects.create_user(username=body['username'], email=body['email'],
                                            password=make_password(body['password']))
            return Response("Success", status=202)
        except IntegrityError:
            return Response("User already exists", status=401)
    else:
        return Response("Passwords don't match", status=401)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rights(request):
    is_editable = request.user.groups.filter(name='Редактирование').exists()
    print(is_editable)
    return JsonResponse({'edit_group': is_editable})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_items(request):
    data = ItemModel.objects.all().values()
    answer = {}
    for element in data:
        answer[element['id']] = element
    print(answer)
    return JsonResponse({'elements': answer})
@require_http_methods(['POST'])
@csrf_exempt
def get_related_items(request):
    data = json.loads(request.body.decode("utf-8"))
    related = ItemRelationModel.objects.filter(to_item=ItemModel.objects.get(id=data['id'])).values()

    item_ids = []
    for element in related:
        item_ids.append(element['from_item_id'])
    data = ItemModel.objects.filter(id__in=item_ids).values()
    return JsonResponse({'elements': list(data)})

@require_http_methods(['POST'])
@csrf_exempt
def sort_rel(request):
    data = json.loads(request.body.decode("utf-8"))
    sort_by = data['attribute']
    sort_to = data['sort']

    related = ItemRelationModel.objects.filter(to_item=ItemModel.objects.get(id=data['id'])).values()

    item_ids = []
    for element in related:
        item_ids.append(element['from_item_id'])

    if (sort_to == 'DESC'):
        data = ItemModel.objects.filter(id__in=item_ids).order_by(sort_by).values()
    else:
        data = ItemModel.objects.filter(id__in=item_ids).order_by(f'-{sort_by}').values()

    return JsonResponse({'elements': list(data)})


@csrf_exempt
def sort(request):
    data = json.loads(request.body.decode("utf-8"))
    sort_by = data['attribute']
    sort_to = data['sort']
    if (sort_to == 'DESC'):
        data = ItemModel.objects.all().order_by(sort_by).values()
    else:
        data = ItemModel.objects.all().order_by(f'-{sort_by}').values()

    return JsonResponse({'elements': list(data)})


@require_http_methods(['POST'])
@csrf_exempt
def save(request):
    data = json.loads(request.body.decode("utf-8"))
    element_id = data['element_id']
    attr_name = data['attribute']
    attr_value = data['value']

    element = ItemModel.objects.get(id=element_id)
    setattr(element, attr_name,
            attr_value)  # приходит один атрибут и его нужно установить динамически, через model.attr не получится
    element.save()

    return JsonResponse({'status': 'ok'})


@require_http_methods(['POST'])
@csrf_exempt
def remove(request):
    data = json.loads(request.body.decode("utf-8"))
    element_id = data['element_id']

    element = ItemModel.objects.get(id=element_id)
    element.delete()

    return JsonResponse({'status': 'ok'})


@require_http_methods(['POST'])
@csrf_exempt
def add(request):  # ItemModel.objects.filter(date_in__gte="2023-01-04", date_out__lte="2023-02-15")
    data = json.loads(request.body.decode("utf-8"))

    element = ItemModel(
        equipment_type=data['equipment_type'],
        manufacturer=data['manufacturer'],
        name=data['name'],
        price=data['price'],
        owner=data['owner'],
        belonging=data['belonging'],
        number=data['number'],
        date_in=data['date_in'],
        date_out=data['date_out'],
    )
    try:
        element.save()
        if ('relatedItem' in data):
            relation = ItemRelationModel(
                from_item=element,
                to_item=ItemModel.objects.get(id=data['relatedItem'])
            )
            relation.save()
    except ValidationError:
        return JsonResponse({'status': 'dateValidationError'})

    return JsonResponse({'status': 'ok'})


@require_http_methods(['POST'])
@csrf_exempt
def excel_export(request):
    data = json.loads(request.body.decode("utf-8"))
    left_date = data['left_date']
    right_date = data['right_date']
    objects = ItemModel.objects.filter(date_in__gte=left_date).order_by('date_in')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Movies'

    columns = [
        'Прибыло/Убыло',
        'Тип',
        'Производитель',
        'Наименование оборудования',
        'Стоимость оборудования',
        'Принадлежность оборудования',
        'Закрепление оборудования за сотрудниками',
        'Накладной номер оборудования',
        'Дата получения оборудования',
        'Дата списания/выдачи оборудования из ремонта',
    ]

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for item in objects:
        row_num += 1

        row = [
            'Прибыло',
            item.equipment_type,
            item.manufacturer,
            item.name,
            item.price,
            item.owner,
            item.belonging,
            item.number,
            item.date_in,
            item.date_out,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    objects = ItemModel.objects.filter(date_out__lte=right_date).order_by('date_out')
    for item in objects:
        if (item.date_out <= datetime.datetime.strptime(right_date,
                                                                                                    "%Y-%m-%d").date()):
            row_num += 1

            row = [
                'Убыло',
                item.equipment_type,
                item.manufacturer,
                item.name,
                item.price,
                item.owner,
                item.belonging,
                item.number,
                item.date_in,
                item.date_out,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
    file_name = f'excel_files/export_{datetime.datetime.now()}.xlsx'
    file_path = settings.BASE_DIR / file_name
    workbook.save(file_path)
    return JsonResponse({'status': 'ok', 'url': str(file_name)})
